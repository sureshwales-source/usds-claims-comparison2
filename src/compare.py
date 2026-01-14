#!/usr/bin/env python3
"""
claims_compare_tool

A small comparison/validation tool for parallel-run claim processing systems.

Design goals:
- Deterministic and reproducible (pure file inputs -> same outputs)
- Transparent mismatch classification (missing/extra/duplicate/value/type)
- Scales to large CSVs via chunked reads + SQLite staging
- Produces an analyst-friendly Excel workbook with drill-down tabs

This repo assumes:
- "old system" outputs are the CMS DE-SynPUF raw files (ground truth baseline)
- "new system" outputs are "comparable outputs" produced by the replacement system.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Dict, Tuple, Optional

import pandas as pd
import numpy as np

# ---------------------------------------------------------------------
# Debugging / validation helper notes
# ---------------------------------------------------------------------
# Throughout this file you'll see lines like:
#     # print("[DEBUG] ...")
# These are intentionally commented out so the tool runs quietly by default.
# If you're validating behavior or troubleshooting a mismatch, simply
# uncomment the prints you care about and re-run.
#
# Tip: Start by uncommenting the prints in `run_compare()` and
# `field_mismatches()` to see high-level progress + mismatch counts.
#
# IMPORTANT: We avoid a global DEBUG flag on purpose—reviewers can see
# exactly what would be printed and you can enable only what you need.
# ---------------------------------------------------------------------


# -----------------------------
# Utilities
# -----------------------------

DATE_COLS = {"BENE_BIRTH_DT", "BENE_DEATH_DT"}
KEY_COL = "DESYNPUF_ID"

PAYMENT_COLS = [
    "MEDREIMB_IP", "BENRES_IP", "PPPYMT_IP",
    "MEDREIMB_OP", "BENRES_OP", "PPPYMT_OP",
    "MEDREIMB_CAR", "BENRES_CAR", "PPPYMT_CAR",
]

DIM_COLS = [
    "BENE_SEX_IDENT_CD", "BENE_RACE_CD", "BENE_ESRD_IND",
    "SP_STATE_CODE",
]


def as_date_yyyymmdd(x) -> Optional[str]:
    """Normalize a YYYYMMDD-ish value to ISO date string (YYYY-MM-DD).
    Returns None when missing/invalid.
    """
    if pd.isna(x):
        return None
    s = str(x).strip()
    # tolerate strings like "1943101" (missing leading zero) by left-padding
    if s.isdigit():
        if len(s) < 8:
            s = s.zfill(8)
        if len(s) == 8:
            try:
                return dt.date(int(s[0:4]), int(s[4:6]), int(s[6:8])).isoformat()
            except ValueError:
                return None
    return None


def standardize_df(df: pd.DataFrame) -> pd.DataFrame:
    """Standardize columns/types enough to compare reliably."""
    df = df.copy()
    # Ensure all expected columns exist (some new outputs may have missing columns)
    for col in DATE_COLS:
        if col in df.columns:
            df[col] = df[col].map(as_date_yyyymmdd)
    # Cast payment columns to numeric (coerce errors to NaN so we can detect format issues)
    for col in PAYMENT_COLS:
        # print(f"[DEBUG]   comparing field: {col}")
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    # Keep dimensions as strings for stable grouping
    for col in DIM_COLS:
        if col in df.columns:
            df[col] = df[col].astype("string")
    if KEY_COL in df.columns:
        df[KEY_COL] = df[KEY_COL].astype("string")
    return df


@dataclass
class TableSpec:
    name: str
    old_path: Path
    new_path: Path
    key_cols: List[str]


def discover_beneficiary_tables(old_dir: Path, new_dir: Path) -> List[TableSpec]:
    """Find matching Beneficiary Summary CSVs in old/new directories."""
    specs: List[TableSpec] = []
    for old_file in sorted(old_dir.glob("DE1_0_*_Beneficiary_Summary_File_Sample_1.csv")):
        new_file = new_dir / old_file.name
        if new_file.exists():
            year = old_file.name.split("_")[2]
            specs.append(TableSpec(
                name=f"beneficiary_{year}",
                old_path=old_file,
                new_path=new_file,
                key_cols=[KEY_COL],
            ))
    return specs


# -----------------------------
# SQLite staging
# -----------------------------

def load_csv_to_sqlite(
    conn: sqlite3.Connection,
    table: str,
    csv_path: Path,
    chunksize: int = 100_000,
) -> Dict[str, float]:
    """Load a CSV into SQLite using chunked pandas reads. Returns load stats."""
    start = dt.datetime.now()
    total_rows = 0
    for i, chunk in enumerate(pd.read_csv(csv_path, chunksize=chunksize)):
        chunk = standardize_df(chunk)
            # print(f"[DEBUG]   chunk {chunk_idx}: standardized cols={len(chunk.columns)}")
        # print(f"[DEBUG]   chunk {chunk_idx}: writing to SQLite table {table_name}")
            chunk.to_sql(table, conn, if_exists="replace" if i == 0 else "append", index=False)
        total_rows += len(chunk)
    elapsed = (dt.datetime.now() - start).total_seconds()
    return {"rows": total_rows, "seconds": elapsed}


def sqlite_query_df(conn: sqlite3.Connection, sql: str, params: tuple = ()) -> pd.DataFrame:
    return pd.read_sql_query(sql, conn, params=params)


# -----------------------------
# Comparison logic
# -----------------------------

def key_health(conn: sqlite3.Connection, old_tbl: str, new_tbl: str, key_col: str) -> pd.DataFrame:
    """Compute missing/extra/duplicate key counts."""
    # Count duplicates
    dup_old = sqlite_query_df(conn, f"""
        SELECT COUNT(*) AS dup_rows
        FROM (
            SELECT {key_col}, COUNT(*) c FROM {old_tbl} GROUP BY {key_col} HAVING c > 1
        )
    """)["dup_rows"].iloc[0]

    dup_new = sqlite_query_df(conn, f"""
        SELECT COUNT(*) AS dup_rows
        FROM (
            SELECT {key_col}, COUNT(*) c FROM {new_tbl} GROUP BY {key_col} HAVING c > 1
        )
    """)["dup_rows"].iloc[0]

    missing_in_new = sqlite_query_df(conn, f"""
        SELECT COUNT(*) AS n
        FROM {old_tbl} o
        LEFT JOIN {new_tbl} n
        ON o.{key_col} = n.{key_col}
        WHERE n.{key_col} IS NULL
    """)["n"].iloc[0]

    extra_in_new = sqlite_query_df(conn, f"""
        SELECT COUNT(*) AS n
        FROM {new_tbl} n
        LEFT JOIN {old_tbl} o
        ON o.{key_col} = n.{key_col}
        WHERE o.{key_col} IS NULL
    """)["n"].iloc[0]

    return pd.DataFrame([{
        "duplicate_keys_old": int(dup_old),
        "duplicate_keys_new": int(dup_new),
        "missing_keys_in_new": int(missing_in_new),
        "extra_keys_in_new": int(extra_in_new),
    }])


def field_mismatches(
    conn: sqlite3.Connection,
    old_tbl: str,
    new_tbl: str,
    key_col: str,
    fields: List[str],
    tolerance: float = 0.0,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Compare numeric fields and return (summary, top_diffs)."""
    summaries = []
    top_rows = []

    for f in fields:
        # Mismatch count (including null asymmetry)
        df = sqlite_query_df(conn, f"""
            SELECT
              SUM(CASE
                    WHEN o.{f} IS NULL AND n.{f} IS NULL THEN 0
                    WHEN o.{f} IS NULL AND n.{f} IS NOT NULL THEN 1
                    WHEN o.{f} IS NOT NULL AND n.{f} IS NULL THEN 1
                    WHEN ABS(o.{f} - n.{f}) > ? THEN 1
                    ELSE 0
                  END) AS mismatch_rows,
              COUNT(*) AS compared_rows,
              SUM(COALESCE(o.{f},0)) AS old_sum,
              SUM(COALESCE(n.{f},0)) AS new_sum,
              SUM(COALESCE(n.{f},0) - COALESCE(o.{f},0)) AS delta_sum,
              SUM(ABS(COALESCE(n.{f},0) - COALESCE(o.{f},0))) AS abs_delta_sum
            FROM {old_tbl} o
            INNER JOIN {new_tbl} n
              ON o.{key_col} = n.{key_col}
        """, (tolerance,))
        row = df.iloc[0].to_dict()
        row["field"] = f
        row["tolerance"] = tolerance
        # percent mismatch
        row["mismatch_rate"] = float(row["mismatch_rows"]) / float(row["compared_rows"]) if row["compared_rows"] else 0.0
        summaries.append(row)

        # Top absolute diffs
        tops = sqlite_query_df(conn, f"""
            SELECT
              o.{key_col} AS {key_col},
              o.{f} AS old_value,
              n.{f} AS new_value,
              (COALESCE(n.{f},0) - COALESCE(o.{f},0)) AS delta,
              ABS(COALESCE(n.{f},0) - COALESCE(o.{f},0)) AS abs_delta
            FROM {old_tbl} o
            INNER JOIN {new_tbl} n
              ON o.{key_col} = n.{key_col}
            WHERE (o.{f} IS NULL AND n.{f} IS NOT NULL)
               OR (o.{f} IS NOT NULL AND n.{f} IS NULL)
               OR ABS(o.{f} - n.{f}) > ?
            ORDER BY abs_delta DESC
            LIMIT 50
        """, (tolerance,))
        if len(tops):
            tops["field"] = f
            top_rows.append(tops)

    summary_df = pd.DataFrame(summaries).sort_values(["mismatch_rate","abs_delta_sum"], ascending=[False, False])
    top_df = pd.concat(top_rows, ignore_index=True) if top_rows else pd.DataFrame()
    return summary_df, top_df


def trend_by_dimension(
    conn: sqlite3.Connection,
    old_tbl: str,
    new_tbl: str,
    key_col: str,
    dim: str,
    fields: List[str],
    tolerance: float,
) -> pd.DataFrame:
    """For each dimension value, compute mismatch counts for any of the given fields."""
    # Build a "mismatch_any" expression
    mismatch_terms = []
    for f in fields:
        mismatch_terms.append(f"""
            (CASE
               WHEN o.{f} IS NULL AND n.{f} IS NULL THEN 0
               WHEN o.{f} IS NULL AND n.{f} IS NOT NULL THEN 1
               WHEN o.{f} IS NOT NULL AND n.{f} IS NULL THEN 1
               WHEN ABS(o.{f} - n.{f}) > {tolerance} THEN 1
               ELSE 0
             END)
        """)
    mismatch_any = " + ".join(mismatch_terms)
    sql = f"""
        SELECT
          COALESCE(o.{dim}, 'UNKNOWN') AS {dim},
          COUNT(*) AS compared_rows,
          SUM(CASE WHEN ({mismatch_any}) > 0 THEN 1 ELSE 0 END) AS any_field_mismatch_rows
        FROM {old_tbl} o
        INNER JOIN {new_tbl} n
          ON o.{key_col} = n.{key_col}
        GROUP BY COALESCE(o.{dim}, 'UNKNOWN')
        ORDER BY any_field_mismatch_rows DESC
    """
    df = sqlite_query_df(conn, sql)
    df["mismatch_rate"] = df["any_field_mismatch_rows"] / df["compared_rows"]
    return df


# -----------------------------
# Reporting
# -----------------------------

def write_excel_report(
    out_path: Path,
    run_meta: Dict,
    overview: pd.DataFrame,
    key_issues: pd.DataFrame,
    mismatch_summary: pd.DataFrame,
    top_diffs: pd.DataFrame,
    trends: Dict[str, pd.DataFrame],
    samples: Dict[str, pd.DataFrame],
) -> None:
    """Write an Excel workbook with multiple tabs + a couple of simple charts."""
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    wb.remove(wb.active)

    def add_sheet(name: str, df: pd.DataFrame):
        ws = wb.create_sheet(title=name[:31])
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        # header style
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        return ws

    # Meta / Overview
    meta_df = pd.DataFrame([run_meta])
    add_sheet("Run_Metadata", meta_df)
    ws_over = add_sheet("Overview", overview)
    add_sheet("Key_Issues", key_issues)
    ws_ms = add_sheet("Payment_Field_Mismatch", mismatch_summary)
    if len(top_diffs):
        add_sheet("Top_Differences", top_diffs)

    for k, df in trends.items():
        add_sheet(f"Trend_{k}", df)

    for k, df in samples.items():
        add_sheet(f"Sample_{k}", df)

    # Add a simple chart to mismatch summary (top 10 mismatch_rate)
    if len(mismatch_summary):
        ws = ws_ms
        # Put a chart below the table
        chart = BarChart()
        chart.title = "Mismatch rate by payment field (top 10)"
        chart.y_axis.title = "Mismatch rate"
        chart.x_axis.title = "Field"
        n = min(10, len(mismatch_summary))
        data = Reference(ws, min_col=ws_over.max_column+1, min_row=1, max_row=1)  # dummy
        # Locate columns by header names
        headers = [c.value for c in ws[1]]
        field_col = headers.index("field") + 1
        rate_col = headers.index("mismatch_rate") + 1
        data = Reference(ws, min_col=rate_col, min_row=1, max_row=n+1)
        cats = Reference(ws, min_col=field_col, min_row=2, max_row=n+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "J2")

    wb.save(out_path)


# -----------------------------
# CLI
# -----------------------------

def run_compare(old_dir: Path, new_dir: Path, out_dir: Path, db_path: Path, tolerance: float = 0.0):
    out_dir.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)

    run_meta = {
        "run_utc": dt.datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "old_dir": str(old_dir),
        "new_dir": str(new_dir),
        "tolerance": tolerance,
    }

    specs = discover_beneficiary_tables(old_dir, new_dir)
    if not specs:
        raise SystemExit(f"No matching Beneficiary Summary CSVs found in {old_dir} and {new_dir}")

    overview_rows = []
    key_issue_rows = []
    mismatch_summaries = []
    top_diffs_all = []
    trends = {}
    samples = {}

    for spec in specs:
        # print(f"[DEBUG] processing table spec: {spec.name}")
        old_tbl = f"{spec.name}_old"
        new_tbl = f"{spec.name}_new"

        load_old = load_csv_to_sqlite(conn, old_tbl, spec.old_path)
        load_new = load_csv_to_sqlite(conn, new_tbl, spec.new_path)

        overview_rows.append({
            "table": spec.name,
            "old_rows": load_old["rows"],
            "new_rows": load_new["rows"],
            "old_load_seconds": load_old["seconds"],
            "new_load_seconds": load_new["seconds"],
        })

        kh = key_health(conn, old_tbl, new_tbl, KEY_COL)
        kh.insert(0, "table", spec.name)
        key_issue_rows.append(kh)

        ms, top = field_mismatches(conn, old_tbl, new_tbl, KEY_COL, PAYMENT_COLS, tolerance=tolerance)
        ms.insert(0, "table", spec.name)
        mismatch_summaries.append(ms)
        if len(top):
            top.insert(0, "table", spec.name)
            top_diffs_all.append(top)

        # Trends
        for dim in ["SP_STATE_CODE", "BENE_RACE_CD"]:
            tr = trend_by_dimension(conn, old_tbl, new_tbl, KEY_COL, dim, PAYMENT_COLS, tolerance=tolerance)
            tr.insert(0, "table", spec.name)
            trends[f"{spec.name}_{dim}"] = tr

        # Small samples: missing keys, duplicate keys in new
        missing = sqlite_query_df(conn, f"""
            SELECT o.{KEY_COL}
            FROM {old_tbl} o
            LEFT JOIN {new_tbl} n ON o.{KEY_COL} = n.{KEY_COL}
            WHERE n.{KEY_COL} IS NULL
            LIMIT 200
        """)
        if len(missing):
            samples[f"{spec.name}_missing_in_new"] = missing

        dup_new = sqlite_query_df(conn, f"""
            SELECT {KEY_COL}, COUNT(*) AS copies
            FROM {new_tbl}
            GROUP BY {KEY_COL}
            HAVING copies > 1
            ORDER BY copies DESC
            LIMIT 200
        """)
        if len(dup_new):
            samples[f"{spec.name}_dup_keys_new"] = dup_new

    overview = pd.DataFrame(overview_rows)
    key_issues = pd.concat(key_issue_rows, ignore_index=True)
    mismatch_summary = pd.concat(mismatch_summaries, ignore_index=True)\
        .sort_values(["mismatch_rate","abs_delta_sum"], ascending=[False, False])
    top_diffs = pd.concat(top_diffs_all, ignore_index=True) if top_diffs_all else pd.DataFrame()

    # Overall roll-ups
    overall = pd.DataFrame([{
        "tables_compared": len(specs),
        "total_old_rows": int(overview["old_rows"].sum()),
        "total_new_rows": int(overview["new_rows"].sum()),
        "total_missing_keys_in_new": int(key_issues["missing_keys_in_new"].sum()),
        "total_extra_keys_in_new": int(key_issues["extra_keys_in_new"].sum()),
        "total_duplicate_keys_new": int(key_issues["duplicate_keys_new"].sum()),
    }])

    report_path = out_dir / "claims_comparison_report.xlsx"
    write_excel_report(
        report_path,
        run_meta=run_meta,
        overview=overall,
        key_issues=key_issues,
        mismatch_summary=mismatch_summary,
        top_diffs=top_diffs,
        trends=trends,
        samples=samples,
    )
    conn.close()
    return report_path


def build_arg_parser():
    p = argparse.ArgumentParser(description="Compare old vs new claim-processing outputs and generate an Excel report.")
    p.add_argument("--old", required=True, type=Path, help="Directory containing baseline (old system) CSV outputs")
    p.add_argument("--new", required=True, type=Path, help="Directory containing candidate (new system) CSV outputs")
    p.add_argument("--out", required=True, type=Path, help="Output directory for reports")
    p.add_argument("--db", default=Path("staging.sqlite"), type=Path, help="SQLite path for staging/comparison")
    p.add_argument("--tolerance", default=0.0, type=float, help="Numeric tolerance for payment comparisons")
    return p


def main():
    args = build_arg_parser().parse_args()
    report = run_compare(args.old, args.new, args.out, args.db, tolerance=args.tolerance)
    print(f"Wrote report: {report}")


if __name__ == "__main__":
    main()
